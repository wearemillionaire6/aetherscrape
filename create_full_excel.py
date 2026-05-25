import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

wb = openpyxl.Workbook()

thin_border = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)

hdr_font   = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
data_font  = Font(name='Calibri', size=10)
phone_font = Font(name='Calibri', size=11, bold=True, color="2471A3")
title_font = Font(name='Calibri', size=16, bold=True, color="1B1F2B")
sub_font   = Font(name='Calibri', size=10, italic=True, color="7F8C8D")

# ALL leads — previous + new with phone numbers
# Format: (Name, Category, Area, Phone, Email, Website, Address, Rating)
all_leads = [
    # ═══════ SALONS — ANDHERI ═══════
    ("Look Good Unisex Salon", "Salon", "Andheri East", "+91 83559 28175", "", "", "Shop no.1, Gokul Plaza, Old Nagardas Rd", "4.9"),
    ("Shanuzz Salon", "Salon", "Andheri West", "+91 91524 18567", "", "shanuzz.com", "Ground Floor, Harshvardhan Chambers", "4.9"),
    ("Moon Studios Family Salon", "Salon", "Andheri East", "+91 90048 32184", "", "moonstudiossalon.in", "Shop C, Harmony Apartments, Krishanlal Marwah Marg", "4.9"),
    ("Ddazzle Unisex Salon", "Salon", "Andheri West", "+91 96536 61417", "", "Instagram only", "Shop 36, Laxmi Plaza", "4.9"),
    ("Glitz Salon Academy", "Salon", "Andheri East", "+91 98926 29888", "", "Instagram only", "Shop No.8, Building No.6, MIDC Central Rd", "4.8"),
    ("FTV Salon", "Salon", "Andheri West", "+91 93211 65417", "", "", "Samarth Vaibhav, 112, Lokhandwala Rd", "4.7"),
    ("THE KIN UNISEX SALON", "Salon", "Andheri West", "+91 70212 70618", "", "", "Society No. D5/55, Mhada Rd, near Hanuman Mandir", "4.8"),
    ("Toni&Guy Four Bungalows", "Salon", "Andheri West", "+91 98673 67722", "", "Linktree only", "Shop No 1, Ashish CHS, Building 24", "4.6"),
    ("Jean-Claude Biguine Salon", "Salon", "Andheri West", "+91 75066 61088", "", "biguineindia.co.in", "1st Floor, Springfield Building, Lokhandwala", "4.8"),
    ("Lakme Salon Mahakali Caves", "Salon", "Andheri East", "+91 91376 73339", "", "lakmesalon.in", "102/A, D Building, Bindra Corporate Centre", "4.6"),
    ("Smitstudio Family Salon", "Salon", "Andheri East", "+91 98192 25222", "", "smitstudio.in", "Shop No 10, Makhwana Road, Crescent Landmark", "4.6"),
    ("Vikas Marwah Hair Salon", "Salon", "Lokhandwala", "+91 98671 26303", "", "vikasmarwah.com", "19A, Sunshine Building, Lokhandwala Market", "4.5"),
    ("Harsha & Rakesh Salon", "Salon", "Lokhandwala", "+91 99203 27888", "", "harsharakesh.com", "Shop No. 6/7, Primerose Apartment, 4th Cross Rd", "4.8"),
    ("Huppa Huiyya Luxury Salon", "Salon", "Veera Desai Rd", "+91 81043 49859", "", "huppahuiyyasalon.online", "Shop No 4, Bhavesha Building, Veera Desai Rd", "4.8"),
    ("Beleesa Salon", "Salon", "Andheri West", "+91 81047 55280", "", "Luzo booking app", "Raviraj Complex, Shop No 11, Rendezvous", "4.9"),
    ("TRESS & BEYOND SALON SPA", "Salon", "Andheri West", "", "", "tressnbeyond.com", "Crystal Plaza, B-08A, New Link Rd, opp. Infinity Mall", "4.3"),
    ("Lakme Salon Azad Nagar", "Salon", "Andheri West", "+91 88791 09088", "", "lakmesalon.in", "Shop No 2, JP Eminence, JP Rd", "4.6"),
    ("Jawed Habib JB Nagar", "Salon", "Andheri East", "+91 83560 82958", "", "jawedhabib.org", "Shop no 1, B wing, Ayushi apartment, near Metro", "4.6"),
    ("Jawed Habib Marol", "Salon", "Andheri East", "+91 74000 68226", "", "jawedhabib.org", "Shop No. 7, Vijay Nagar, Gulmohar Apt, Marol Maroshi Rd", "4.5"),
    ("Awesome Unisex Salon", "Salon", "Andheri", "+91 98203 29444", "", "", "Shop No. 5, Shiv Ashish, Swami Vivekanand Rd", "4.2"),
    ("Blossom Salon", "Salon", "Andheri West", "", "", "", "Shop No 4&5, Nyaja Deep Soc, A Wing, behind SAB TV Tower", "4.7"),

    # ═══════ SALONS — BANDRA ═══════
    ("The Glam Lab Salon", "Salon", "Bandra West", "+91 73040 14055", "", "theglamlabsalon.in", "Shop No. 101, Asra Building, Waterfield Road", "4.9"),
    ("Apple Unisex Salon", "Salon", "Bandra West", "+91 90821 83323", "", "applesalon.in", "Ground Floor, SVD House, 32nd Rd", "4.9"),
    ("7 Wonders Studio", "Salon", "Bandra West", "+91 98336 79949", "", "7wondersstudiosalon.in", "shop no 4, Libra Tower, Hill Rd", "4.9"),
    ("Color Cafe Salon", "Salon", "Bandra West", "+91 80801 26688", "", "colorcafe.co.in", "Kailash Building, 156, Waterfield Road", "4.9"),
    ("Vipul Chudasama Salon", "Salon", "Bandra (Pali Hill)", "", "", "", "Jasmin House First Floor, Pali Mala Road", "4.9"),
    ("Ritu's Hair & Beauty", "Salon", "Bandra West", "", "", "", "shop no 4, Rizvi Mahal, Waterfield Road", "4.9"),
    ("Dessange Salon & Spa", "Salon", "Bandra West", "+91 73043 38957", "", "dessangemumbai.com", "190, Savanna Court, Turner Rd", "4.5"),
    ("Lemon Salons Bandra", "Salon", "Bandra West", "+91 77389 15206", "", "Dingg booking", "Shop No. 4, Sunrise Co. Hsg. Soc, Waterfield Road", "4.6"),
    ("Ccoral Salon", "Salon", "Bandra West", "+91 73044 42223", "", "ccoralsalon.in", "Ground floor, Raut Sadan, St John Baptist Rd", "4.6"),
    ("Jean Claude Olivier Salon", "Salon", "Bandra West", "+91 99877 97773", "", "jeanclaudeolivier.in", "Ground floor, Coral Apartment, Rd Number 24", "4.8"),
    ("Florian Hurel Hair Couture", "Salon", "Bandra West", "", "", "", "Ground floor, Antoine villa, 33, Sherly Mala road", "4.8"),
    ("Lakme Salon Hill Road", "Salon", "Bandra West", "+91 88790 44912", "", "lakmesalon.in", "No C/5, Libra Tower, Hill Rd", "4.8"),
    ("Enrich Salon Bandra", "Salon", "Bandra West", "+91 96198 52607", "", "enrichbeauty.com", "Soney Apartment, 45, Hill Rd, opposite Elco Market", "4.5"),
    ("Lakme Salon Linking Road", "Salon", "Bandra West", "+91 88790 52611", "", "lakmesalon.in", "No 227, Diamond Link, off Linking Road", "4.7"),

    # ═══════ SALONS — MALAD ═══════
    ("De Bella & Beau Salon", "Salon", "Malad West", "+91 98204 22003", "", "WhatsApp only", "Shop 10, Old Sonal Industrial Estate, New Link Rd", "4.8"),
    ("Locks & Lustre Salon", "Salon", "Malad West", "+91 98334 16565", "", "locksandlustresalon.com", "Shop no G-3, Sej Plaza, Marve Rd", "4.8"),
    ("FLOW HAIR STUDIO", "Salon", "Malad West", "+91 89766 50250", "", "flowhairstudio.in", "Shop no. 14, Atlanta Building", "4.9"),
    ("Stylorria Salon", "Salon", "Malad West", "+91 84519 58888", "", "Google Sites", "Shop No 4, Deep Sunder Lane CHS", "4.8"),
    ("Lakme Salon Chincholi", "Salon", "Malad West", "+91 70390 03286", "", "lakmesalon.in", "Bldg No 3, Sony Cplx, Chincholi Bunder Rd", "4.8"),
    ("Alayna Salon", "Salon", "Malad West", "+91 98196 51099", "", "alaynasalon.com", "Shop no 2C, Kiran Tower, off New Link Rd", "4.7"),
    ("Hair Canvas Salon", "Salon", "Malad West", "+91 77180 53483", "", "", "Shop No 3&4, Solitaire Building, New Link Rd", "4.7"),
    ("Enrich Salon Malad", "Salon", "Malad West", "+91 88792 23004", "", "enrichbeauty.com", "Shop No.34/35, 1st Floor, Inorbit Mall", "4.8"),
    ("SHRUSHTI'S Salon", "Salon", "Malad West", "+91 80801 80883", "", "", "Shop no. 4, Vrindavan 1-C CHS Ltd", "4.8"),
    ("INFINITY UNISEX SALON", "Salon", "Malad West", "+91 90499 06565", "", "", "BJ Patel Rd, opp. Raj Computer", "4.7"),
    ("THRIVE UNISEX SALOON", "Salon", "Malad West", "", "", "", "Shop no.6, Shah Arcade-2, Rani Sati Rd", "4.8"),
    ("Just Trim Salon", "Salon", "Malad West", "", "", "", "Shop no.5, Dheeraj Platinum, Chincholi Bunder Rd", "4.8"),

    # ═══════ SALONS — BORIVALI ═══════
    ("Urbane Salon", "Salon", "Borivali West", "+91 91522 01957", "", "", "Shop no 2, Ekta Elegance, Link Road", "4.9"),
    ("Lookout Salon", "Salon", "Borivali West", "+91 98194 02381", "", "lookoutsalon.co.in", "Metro Pillar 247, Sun Sumit Enclave, New Link Rd", "4.8"),
    ("Ecoluxe Salon", "Salon", "Borivali West", "+91 86579 88918", "", "Instagram only", "Shop No 16, Satra Park, Shimpoli Rd", "4.8"),
    ("KS Beauty Lounge", "Salon", "Borivali West", "+91 91525 19145", "", "", "Shop No. 1-2, Kapoor Apartment, Chandavarkar Rd", "4.8"),
    ("Posh Salon & Spa", "Salon", "Borivali West", "+91 74003 74644", "", "", "Vastu Pinnacle, Eksar Rd", "4.8"),
    ("Lakme Salon SV Road Borivali", "Salon", "Borivali West", "+91 93722 46874", "", "lakmesalon.in", "Shop No 5, 6/7, Swami Vivekanand Rd", "4.7"),
    ("Kapil's Salon Borivali", "Salon", "Borivali West", "+91 86557 75644", "", "kapilssalon.com", "Shop No 42, Satra Park, Shimpoli Rd", "4.7"),
    ("Enrich Salon Satra Park", "Salon", "Borivali West", "+91 75065 75768", "", "enrichbeauty.com", "Shop No. 33 & 34, Satra Park, Shimpoli Rd", "4.7"),
    ("Lemon Salons Borivali", "Salon", "Borivali West", "+91 73044 53967", "", "Dingg booking", "Shop no 39, Satra Park, Shimpoli Rd", "4.5"),
    ("Enrich Salon Factory Lane", "Salon", "Borivali West", "+91 96198 52610", "", "enrichbeauty.com", "Shop No 12, Kent Garden Apartment, Factory Ln", "4.7"),
    ("Elinor Salon", "Salon", "Borivali West", "+91 93216 50585", "", "poonamgohel.in", "Shop no 001, Gokul Gagan Apartment, Factory Ln", "4.8"),
    ("LYNN SALON", "Salon", "Borivali West", "", "", "", "Acropolis 1, Laxman Mhatre Rd", "4.8"),
    ("Rac's Salon Borivali", "Salon", "Borivali West", "", "", "", "Shop No. 13, Kamla Rajesh Tower, Chandavarkar Rd", "4.9"),
    ("Elinor Salon", "Salon", "Borivali West", "", "", "", "Shop no 001, Gokul Gagan Apartment, Factory Ln", "4.8"),

    # ═══════ GYMS — ANDHERI ═══════
    ("Waves Gym", "Gym", "Andheri West", "+91 22 6678 7970", "", "wavesgym.com", "5th Floor, Morya Estate, New Link Rd", "4.9"),
    ("IH Remind Fitness", "Gym", "Andheri East", "+91 79898 73329", "", "ihremindfitness.com", "Shop 101, Empress Business Bay, MIDC Central Rd", "4.7"),
    ("Plus Fitness 24/7", "Gym", "Andheri West", "+91 90820 92919", "", "plusfitness.co.in", "Salve Marg, 12A, Chandak Unicorn, Off Veera Desai Rd", "4.8"),
    ("Powerr Vibes Gym", "Gym", "Andheri", "+91 81040 04857", "", "Cult.fit listing", "Basement, Navkar Complex, opp. Andheri Court", "4.8"),
    ("GOODLIFE FITNESS", "Gym", "Andheri", "+91 72086 86233", "", "Instagram only", "Shop no 7, C Wing, Atul, Military Rd", "4.7"),
    ("i5 Fitness", "Gym", "Andheri East", "+91 97691 55315", "", "", "3rd Floor, T5, Pinnacle Business Park, Mahakali Caves Rd", "4.7"),
    ("Zion Fitness", "Gym", "Andheri East", "+91 77382 58050", "", "zionfitness.in", "H.E.K Compound, Mahakali Caves Rd, near Canossa School", "4.6"),
    ("Cloud 9 Fitness Club", "Gym", "Andheri West", "", "", "", "Paramount Bldg, New Link Rd, behind Kohinoor", "4.3"),
    ("Planet Muscle Fitness", "Gym", "Andheri West", "", "", "", "Office 103, Ameya House, near Metro Azad Nagar", "4.8"),
    ("FORTUNE FITNESS", "Gym", "Andheri East", "+91 86929 60009", "", "fortunefitnessgym.com", "Nirvan Corporate, Pumphouse, opp. Aghadi Nagar", "4.4"),
    ("Fitness Hub", "Gym", "Andheri", "", "", "", "Shop 1/2, Ground Floor, Teli Gali, opp. Criticare Hospital", "4.6"),
    ("Fitness Factory", "Gym", "Andheri", "+91 96196 15000", "", "", "4th Floor, Idea Square, above Kalyan Jewellers, opp. Citi Mall", "4.1"),
    ("Stay Fit Andheri", "Gym", "Andheri West", "+91 77978 89788", "info@ironhouse.co.in", "ironhouse.co.in", "Kashibai Mahadev Marg, near The Club", "4.7"),
    ("FITNESS ART", "Gym", "Andheri", "+91 84229 96655", "", "", "1st floor, Deep Society, JP Rd, above DMart", "4.7"),

    # ═══════ SPAS — LOWER PAREL ═══════
    ("Medis Luxury Spa", "Spa", "Lower Parel", "+91 83692 38697", "", "medisluxuryspaparel.in", "Shop 101, Raghuvanshi Mills, Senapati Bapat Marg", "5.0"),
    ("Nature Wellness Spa", "Spa", "Lower Parel", "+91 85917 62214", "", "naturespalowerparel.in", "3rd Floor, Tree Building, Raghuvanshi Mill Compound", "4.9"),
    ("Kyra Luxury Spa", "Spa", "Lower Parel", "+91 93728 31214", "", "kyraspalowerparel.com", "Shop 101, Times Tower, Kamla Mill Compound", "4.8"),
    ("Oceanic Spa", "Spa", "Lower Parel", "+91 81089 52829", "", "oceanicspalowerparel.in", "1st floor, Todi Mill Compound, Senapati Bapat Marg", "4.8"),
    ("Maitri Thai Spa", "Spa", "Lower Parel", "", "", "maitrispaparel.in", "Office No 101, Times Tower, Kamala Mills", "4.9"),
    ("Mudram Wellness Spa", "Spa", "Lower Parel", "", "", "", "1st Floor, Sumandar House, Raghuvanshi Mill Compound", "4.8"),
    ("Madisyn Spa", "Spa", "Lower Parel", "+91 90822 12203", "", "madisynspalowerparel.in", "1st Floor, Sumandar House, Raghuvanshi Mills", "4.8"),
    ("Mudraa Spa", "Spa", "Lower Parel", "+91 96196 08043", "", "mudraaspa.in", "Shop 102, Raghuvanshi Mills, Senapati Bapat Marg", "4.5"),
    ("Yuan Thai Spa", "Spa", "Lower Parel", "+91 98672 00557", "", "yuanthaispa.com", "1st Floor, Raghuvanshi Mills, Senapati Bapat Marg", "4.1"),
    ("R Spa", "Spa", "Lower Parel", "+91 89765 79019", "", "", "Ground Floor, Times Tower, Kamala Mills Compound", "4.8"),
    ("Spring Spa Worli", "Spa", "Worli", "+91 91371 06224", "", "", "Shop 209, 2nd floor, Kachwala Building, Dr Annie Besant Rd", "4.8"),
    ("Richy Spa", "Spa", "Byculla", "+91 91676 14119", "", "richyspachinchpokli.com", "Office No. 6A-6C, Anand Estate, Sane Guruji Marg", "4.9"),
    ("Bodhi Luxury Spa", "Spa", "Lower Parel", "+91 96199 20630", "", "bodhigroup.in", "Store LG22, Phoenix Palladium Mall, Senapati Bapat Marg", "4.1"),
    ("St. Regis Spa", "Spa", "Lower Parel", "+91 88799 78888", "", "marriott.com", "Wellness Floor, Level 10, 462, Senapati Bapat Marg", "4.4"),

    # ═══════ YOGA — BORIVALI ═══════
    ("Yog Shala Tat-Tvam-Asi", "Yoga", "Borivali", "+91 89762 15478", "", "", "Shop No. 2&3, Rajeshwari Building, near Sai Baba Nagar", "5.0"),
    ("Ashtanga Yog Shala", "Yoga", "Borivali", "+91 98923 59792", "", "", "1st floor, The Zone Mall, Chandavarkar Rd", "4.9"),
    ("Yoga365", "Yoga", "Borivali", "+91 98215 23064", "", "yoga365.in", "1st Floor, Plot no. 368, Besides Muthoot Finance", "4.9"),
    ("SoulYog Studio", "Yoga", "Borivali", "+91 83690 72930", "", "WhatsApp only", "Play and Learn Building, near Kanakia Aroha", "5.0"),
    ("Yog Kutir", "Yoga", "Borivali", "+91 92213 13717", "", "yogkutir.in", "1, Adinath Towers, Kanti Park Rd, Shimpoli", "4.8"),
    ("Rhythm & Flow Fitness", "Yoga", "Borivali", "+91 93269 47889", "", "rhythmandflow.in", "Abhiram CHSL, Amarkant Jha Marg, off Shimpoli Road", "4.9"),
    ("Yoga with Fennil", "Yoga", "Borivali", "", "", "", "C-1, Hariom Apartment, Swami Vivekanand Rd", "4.9"),
    ("Iyengar Yoga Pratishthan", "Yoga", "Borivali", "+91 98733 94133", "", "", "64, D1, Dwakesh Park, above HDFC Bank", "4.9"),
    ("Flow & Sway Fitness", "Yoga", "Borivali", "+91 87793 79503", "", "", "Clover Grove, Shop 19 & 20, Chikoowadi", "4.9"),
    ("SHAPE IN YOGA", "Yoga", "Borivali", "+91 77383 00448", "", "", "27, 383-403, RSC Rd Number 26, near St. Rock's School", "5.0"),
    ("OM YOGA STUDIO", "Yoga", "Borivali", "+91 88796 04929", "", "", "Tata Power House, RSM Gym, Phulpakharu Garden", "4.9"),

    # ═══════ PET GROOMING ═══════
    ("Paw Lovey Spa", "Pet Grooming", "Andheri", "+91 75068 80515", "", "", "Shop no.11, Takshila 29B, Mahakali Caves Rd", "5.0"),
    ("Cuddle Grooms", "Pet Grooming", "Mumbai", "+91 85911 89518", "", "", "Shop no.22, Pink Corner", "5.0"),
    ("Pampurr Pets Grooming", "Pet Grooming", "Andheri", "+91 98193 99281", "", "", "Shop No A-2, Crystal Colony, near Canossa High School", "4.9"),
    ("Furry Foots Pet Grooming", "Pet Grooming", "Magathane", "+91 91369 30520", "", "", "Vasant Marvel Complex, near Magathane Metro", "4.9"),
    ("We Love Pets Grooming", "Pet Grooming", "Mumbai", "+91 83695 99051", "", "", "Shop No 9, Nilkhanth Soc.", "4.8"),
    ("Zane's Pet Spa", "Pet Grooming", "Lower Parel", "+91 91524 92637", "", "", "1st Floor, Todi Mill Compound, Senapati Bapat Marg", "4.8"),
    ("FUR N TAILS", "Pet Grooming", "Mumbai", "+91 70211 56081", "", "", "Shop No 2, Vasuki Building, near Somaiya College", "4.7"),
    ("The Happy Tails", "Pet Grooming", "Mumbai", "+91 77158 88881", "", "", "Prabha Niketan, Plot No. 338", "4.6"),

    # ═══════ PHYSIOTHERAPY ═══════
    ("Movmed Physiotherapy", "Physio", "Andheri West", "080 4810 4585", "", "", "Andheri West, Mumbai 400053", "7.3"),
    ("Dr Sahil's Physiotherapy", "Physio", "Thane West", "+91 90297 31346", "", "advphysiotherapyclinic.com", "Shop 13, Bldg 3, Siddhachal Phase 8, Pokharan Rd", "6.7"),
    ("Inest Family Clinic", "Physio", "Panvel", "+91 79779 08975", "", "vedantaaclinic.com", "Shop no 2, Govind Nilayam, opp. Status hotel", "7.1"),
    ("EVOLUTION OF MOVEMENT", "Physio", "Colaba", "080 4812 1136", "", "", "Colaba, 400005", "3.9"),

    # ═══════ YOGA (SULEKHA) ═══════
    ("Shushant Dhuri Yoga", "Yoga", "Vile Parle East", "080 6987 4345", "", "", "Vile Parle East, 400057", "5.6"),
    ("K Fitness & Yoga Institute", "Yoga", "Ghatkopar West", "080 3542 8557", "", "", "Ghatkopar West, 400084", "4.9"),
    ("Nivedita Dhage Yoga", "Yoga", "Dadar West", "080 6983 1607", "", "", "Dadar West, 400028", "6.7"),

    # ═══════ DRIVING SCHOOLS ═══════
    ("Superdrive Zone", "Driving School", "Kandivali West", "080 4810 2575", "", "", "Kandivali West, 400067", "5.0"),
    ("Vaidya Motor Training", "Driving School", "Girgaon", "+91 98216 27774", "", "", "Khotachiwadi Rd, Charni Road East, Girgaon", "5.0"),
    ("Toyota Driving School", "Driving School", "Malad West", "+91 77100 33877", "", "toyotabharat.com", "504, New Link Rd, opp. Goregaon Sports Club", "5.0"),
]

# ═══════ SHEET 1: ALL LEADS ═══════
ws = wb.active
ws.title = "All Leads (Complete)"
ws.sheet_properties.tabColor = "2C3E50"

HEADERS = ["#", "Business Name", "Category", "Area", "📞 Phone", "📧 Email", "🌐 Website", "📍 Address", "⭐ Rating"]
COL_WIDTHS = [4, 30, 15, 18, 22, 25, 28, 50, 8]
HEADER_COLOR = "2C3E50"

# Title
ws.merge_cells("A1:I1")
ws["A1"].value = f"📊 Mumbai Business Leads — Complete Database — {datetime.now().strftime('%d %b %Y')}"
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:I2")
ws["A2"].value = f"Total: {len(all_leads)} leads  |  Source: Google Maps via Firecrawl API  |  Categories: Salons, Gyms, Spas, Yoga, Pet Grooming, Physio, Driving Schools"
ws["A2"].font = sub_font
ws.row_dimensions[2].height = 20

# Headers
hdr_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border
ws.row_dimensions[4].height = 28
ws.freeze_panes = "A5"

# Category colors
cat_colors = {
    "Salon": "EBF5FB",
    "Gym": "E8F8F5",
    "Spa": "F5EEF8",
    "Yoga": "FEF9E7",
    "Pet Grooming": "FDEDEC",
    "Physio": "F0F3F4",
    "Driving School": "FBEEE6",
}

for i, lead in enumerate(all_leads):
    row = i + 5
    name, cat, area, phone, email, website, addr, rating = lead
    vals = [i+1, name, cat, area, phone, email, website, addr, rating]
    fill_color = cat_colors.get(cat, "FFFFFF")
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = phone_font if col == 5 else data_font
        c.fill = row_fill
        c.border = thin_border
        c.alignment = Alignment(vertical='center', wrap_text=True,
                               horizontal='center' if col in (1, 9) else 'left')
    ws.row_dimensions[row].height = 22

for idx, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(idx)].width = w

# Auto-filter
ws.auto_filter.ref = f"A4:I{4 + len(all_leads)}"

# ═══════ SHEET 2: LEADS WITH PHONES ═══════
ws2 = wb.create_sheet("📞 With Phone Numbers")
ws2.sheet_properties.tabColor = "27AE60"

phone_leads = [l for l in all_leads if l[3] and l[3].strip()]

ws2.merge_cells("A1:I1")
ws2["A1"].value = f"📞 Leads With Verified Phone Numbers — {len(phone_leads)} contacts ready for cold calling"
ws2["A1"].font = Font(name='Calibri', size=14, bold=True, color="27AE60")
ws2.row_dimensions[1].height = 32

for col, h in enumerate(HEADERS, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = hdr_font
    c.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border
ws2.freeze_panes = "A4"

for i, lead in enumerate(phone_leads):
    row = i + 4
    name, cat, area, phone, email, website, addr, rating = lead
    vals = [i+1, name, cat, area, phone, email, website, addr, rating]
    fill_color = cat_colors.get(cat, "FFFFFF")
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=row, column=col, value=val)
        c.font = phone_font if col == 5 else data_font
        c.fill = row_fill
        c.border = thin_border
        c.alignment = Alignment(vertical='center', horizontal='center' if col in (1, 9) else 'left')
    ws2.row_dimensions[row].height = 22

for idx, w in enumerate(COL_WIDTHS, 1):
    ws2.column_dimensions[get_column_letter(idx)].width = w
ws2.auto_filter.ref = f"A3:I{3 + len(phone_leads)}"

# ═══════ SHEET 3: CALL TRACKER ═══════
ws3 = wb.create_sheet("📋 Call Tracker")
ws3.sheet_properties.tabColor = "E74C3C"

ws3.merge_cells("A1:J1")
ws3["A1"].value = "📋 Cold Call Tracker — Log your outreach"
ws3["A1"].font = Font(name='Calibri', size=14, bold=True, color="E74C3C")
ws3.row_dimensions[1].height = 32

tracker_headers = ["#", "Business Name", "Category", "Area", "📞 Phone", "Call Date", "Spoke To", "Outcome", "Follow-Up", "Notes"]
tracker_widths = [4, 28, 13, 16, 22, 13, 18, 18, 13, 35]

for col, h in enumerate(tracker_headers, 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = hdr_font
    c.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border
ws3.freeze_panes = "A4"

for i, lead in enumerate(phone_leads):
    row = i + 4
    name, cat, area, phone = lead[0], lead[1], lead[2], lead[3]
    vals = [i+1, name, cat, area, phone, "", "", "", "", ""]
    for col, val in enumerate(vals, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = phone_font if col == 5 else data_font
        c.border = thin_border
        c.alignment = Alignment(vertical='center')
        if i % 2 == 0:
            c.fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type='solid')
    ws3.row_dimensions[row].height = 22

for idx, w in enumerate(tracker_widths, 1):
    ws3.column_dimensions[get_column_letter(idx)].width = w

dv = DataValidation(type="list", formula1='"Not Answered,Interested,Not Interested,Call Back,Wrong Number,Meeting Set,Closed Deal"')
ws3.add_data_validation(dv)
for row in range(4, 4 + len(phone_leads)):
    dv.add(ws3.cell(row=row, column=8))

# ═══════ SHEET 4: BY CATEGORY ═══════
for cat_name, tab_color in [("Salons", "3498DB"), ("Gyms", "2ECC71"), ("Spas", "9B59B6"), ("Yoga", "F1C40F"), ("Pet Grooming", "E74C3C")]:
    cat_leads = [l for l in all_leads if l[1] == cat_name.rstrip('s') or l[1] == cat_name]
    if cat_name == "Salons":
        cat_leads = [l for l in all_leads if l[1] == "Salon"]
    elif cat_name == "Gyms":
        cat_leads = [l for l in all_leads if l[1] == "Gym"]
    elif cat_name == "Spas":
        cat_leads = [l for l in all_leads if l[1] == "Spa"]
    
    if not cat_leads:
        continue
    
    ws_cat = wb.create_sheet(f"📁 {cat_name}")
    ws_cat.sheet_properties.tabColor = tab_color
    
    ws_cat.merge_cells("A1:I1")
    ws_cat["A1"].value = f"{cat_name} — {len(cat_leads)} leads"
    ws_cat["A1"].font = Font(name='Calibri', size=14, bold=True, color=tab_color)
    ws_cat.row_dimensions[1].height = 32
    
    for col, h in enumerate(HEADERS, 1):
        c = ws_cat.cell(row=3, column=col, value=h)
        c.font = hdr_font
        c.fill = PatternFill(start_color=tab_color, end_color=tab_color, fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border
    ws_cat.freeze_panes = "A4"
    
    for i, lead in enumerate(cat_leads):
        row = i + 4
        vals = [i+1] + list(lead)
        for col, val in enumerate(vals, 1):
            c = ws_cat.cell(row=row, column=col, value=val)
            c.font = phone_font if col == 5 else data_font
            c.border = thin_border
            c.alignment = Alignment(vertical='center', horizontal='center' if col in (1, 9) else 'left')
        ws_cat.row_dimensions[row].height = 22
    
    for idx, w in enumerate(COL_WIDTHS, 1):
        ws_cat.column_dimensions[get_column_letter(idx)].width = w

# ═══════ SAVE ═══════
path = "/Users/bhaveshashokwaghmare/Desktop/anticlaude/Mumbai_All_Leads_Complete.xlsx"
wb.save(path)
print(f"✅ Excel saved: {path}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Total leads: {len(all_leads)}")
print(f"   With phones: {len(phone_leads)}")
