import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mumbai Leads"

# ── Simple clean styling ──
thin = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
header_font = Font(name='Arial', bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type='solid')
data_font = Font(name='Arial', size=11)
phone_font = Font(name='Arial', size=11, bold=True)
alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type='solid')
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')

# ── ALL LEADS in one flat list ──
leads = [
    # (Name, Category, Area, Phone, Email, Website, Address, Rating)
    # SALONS — ANDHERI
    ("Look Good Unisex Salon", "Salon", "Andheri East", "+91 83559 28175", "", "", "Shop no.1, Gokul Plaza, Old Nagardas Rd", "4.9"),
    ("Shanuzz Salon", "Salon", "Andheri West", "+91 91524 18567", "", "shanuzz.com", "Ground Floor, Harshvardhan Chambers", "4.9"),
    ("Moon Studios Family Salon", "Salon", "Andheri East", "+91 90048 32184", "", "moonstudiossalon.in", "Shop C, Harmony Apartments, Krishanlal Marwah Marg", "4.9"),
    ("Ddazzle Unisex Salon", "Salon", "Andheri West", "+91 96536 61417", "", "Instagram only", "Shop 36, Laxmi Plaza", "4.9"),
    ("Glitz Salon Academy", "Salon", "Andheri East", "+91 98926 29888", "", "Instagram only", "Shop No.8, Building No.6, MIDC Central Rd", "4.8"),
    ("FTV Salon", "Salon", "Andheri West", "+91 93211 65417", "", "", "Samarth Vaibhav, 112, Lokhandwala Rd", "4.7"),
    ("THE KIN UNISEX SALON", "Salon", "Andheri West", "+91 70212 70618", "", "", "Society No. D5/55, Mhada Rd", "4.8"),
    ("Toni&Guy Four Bungalows", "Salon", "Andheri West", "+91 98673 67722", "", "Linktree only", "Shop No 1, Ashish CHS, Building 24", "4.6"),
    ("Jean-Claude Biguine Salon", "Salon", "Andheri West", "+91 75066 61088", "", "biguineindia.co.in", "1st Floor, Springfield Building, Lokhandwala", "4.8"),
    ("Lakme Salon Mahakali Caves", "Salon", "Andheri East", "+91 91376 73339", "", "lakmesalon.in", "102/A, D Building, Bindra Corporate Centre", "4.6"),
    ("Smitstudio Family Salon", "Salon", "Andheri East", "+91 98192 25222", "", "smitstudio.in", "Shop No 10, Makhwana Road, Crescent Landmark", "4.6"),
    ("Vikas Marwah Hair Salon", "Salon", "Lokhandwala", "+91 98671 26303", "", "vikasmarwah.com", "19A, Sunshine Building, Lokhandwala Market", "4.5"),
    ("Harsha & Rakesh Salon", "Salon", "Lokhandwala", "+91 99203 27888", "", "harsharakesh.com", "Shop No. 6/7, Primerose Apartment, 4th Cross Rd", "4.8"),
    ("Huppa Huiyya Luxury Salon", "Salon", "Andheri West", "+91 81043 49859", "", "huppahuiyyasalon.online", "Shop No 4, Bhavesha Building, Veera Desai Rd", "4.8"),
    ("Beleesa Salon", "Salon", "Andheri West", "+91 81047 55280", "", "Luzo booking", "Raviraj Complex, Shop No 11, Rendezvous", "4.9"),
    ("Lakme Salon Azad Nagar", "Salon", "Andheri West", "+91 88791 09088", "", "lakmesalon.in", "Shop No 2, JP Eminence, JP Rd", "4.6"),
    ("Jawed Habib JB Nagar", "Salon", "Andheri East", "+91 83560 82958", "", "jawedhabib.org", "Shop no 1, B wing, Ayushi apartment, near Metro", "4.6"),
    ("Jawed Habib Marol", "Salon", "Andheri East", "+91 74000 68226", "", "jawedhabib.org", "Shop No. 7, Vijay Nagar, Gulmohar Apt", "4.5"),
    ("Awesome Unisex Salon", "Salon", "Andheri", "+91 98203 29444", "", "", "Shop No. 5, Shiv Ashish, Swami Vivekanand Rd", "4.2"),
    ("Blossom Salon", "Salon", "Andheri West", "", "", "", "Shop No 4&5, Nyaja Deep Soc, behind SAB TV Tower", "4.7"),
    ("TRESS & BEYOND SALON SPA", "Salon", "Andheri West", "", "", "tressnbeyond.com", "Crystal Plaza, B-08A, New Link Rd", "4.3"),

    # SALONS — BANDRA
    ("The Glam Lab Salon", "Salon", "Bandra West", "+91 73040 14055", "", "theglamlabsalon.in", "Shop No. 101, Asra Building, Waterfield Road", "4.9"),
    ("Apple Unisex Salon", "Salon", "Bandra West", "+91 90821 83323", "", "applesalon.in", "Ground Floor, SVD House, 32nd Rd", "4.9"),
    ("7 Wonders Studio", "Salon", "Bandra West", "+91 98336 79949", "", "7wondersstudiosalon.in", "shop no 4, Libra Tower, Hill Rd", "4.9"),
    ("Color Cafe Salon", "Salon", "Bandra West", "+91 80801 26688", "", "colorcafe.co.in", "Kailash Building, 156, Waterfield Road", "4.9"),
    ("Vipul Chudasama Salon", "Salon", "Bandra Pali Hill", "", "", "", "Jasmin House First Floor, Pali Mala Road", "4.9"),
    ("Ritu's Hair & Beauty", "Salon", "Bandra West", "", "", "", "shop no 4, Rizvi Mahal, Waterfield Road", "4.9"),
    ("Dessange Salon & Spa", "Salon", "Bandra West", "+91 73043 38957", "", "dessangemumbai.com", "190, Savanna Court, Turner Rd", "4.5"),
    ("Lemon Salons Bandra", "Salon", "Bandra West", "+91 77389 15206", "", "Dingg booking", "Shop No. 4, Sunrise Co. Hsg. Soc, Waterfield Road", "4.6"),
    ("Ccoral Salon", "Salon", "Bandra West", "+91 73044 42223", "", "ccoralsalon.in", "Ground floor, Raut Sadan, St John Baptist Rd", "4.6"),
    ("Jean Claude Olivier Salon", "Salon", "Bandra West", "+91 99877 97773", "", "jeanclaudeolivier.in", "Ground floor, Coral Apartment, Rd Number 24", "4.8"),
    ("Florian Hurel Hair Couture", "Salon", "Bandra West", "", "", "", "Ground floor, Antoine villa, 33, Sherly Mala road", "4.8"),
    ("Lakme Salon Hill Road", "Salon", "Bandra West", "+91 88790 44912", "", "lakmesalon.in", "No C/5, Libra Tower, Hill Rd", "4.8"),
    ("Enrich Salon Bandra", "Salon", "Bandra West", "+91 96198 52607", "", "enrichbeauty.com", "Soney Apartment, 45, Hill Rd, opposite Elco Market", "4.5"),
    ("Lakme Salon Linking Road", "Salon", "Bandra West", "+91 88790 52611", "", "lakmesalon.in", "No 227, Diamond Link, off Linking Road", "4.7"),

    # SALONS — MALAD
    ("De Bella & Beau Salon", "Salon", "Malad West", "+91 98204 22003", "", "WhatsApp only", "Shop 10, Old Sonal Industrial Estate, New Link Rd", "4.8"),
    ("Locks & Lustre Salon", "Salon", "Malad West", "+91 98334 16565", "", "locksandlustresalon.com", "Shop no G-3, Sej Plaza, Marve Rd", "4.8"),
    ("FLOW HAIR STUDIO", "Salon", "Malad West", "+91 89766 50250", "", "flowhairstudio.in", "Shop no. 14, Atlanta Building", "4.9"),
    ("Stylorria Salon", "Salon", "Malad West", "+91 84519 58888", "", "Google Sites", "Shop No 4, Deep Sunder Lane CHS", "4.8"),
    ("Lakme Salon Chincholi", "Salon", "Malad West", "+91 70390 03286", "", "lakmesalon.in", "Bldg No 3, Sony Cplx, Chincholi Bunder Rd", "4.8"),
    ("Alayna Salon", "Salon", "Malad West", "+91 98196 51099", "", "alaynasalon.com", "Shop no 2C, Kiran Tower, off New Link Rd", "4.7"),
    ("Hair Canvas Salon", "Salon", "Malad West", "+91 77180 53483", "", "", "Shop No 3&4, Solitaire Building, New Link Rd", "4.7"),
    ("Enrich Salon Malad", "Salon", "Malad West", "+91 88792 23004", "", "enrichbeauty.com", "Shop No.34/35, 1st Floor, Inorbit Mall", "4.8"),
    ("SHRUSHTI'S Salon", "Salon", "Malad West", "+91 80801 80883", "", "", "Shop no. 4, Vrindavan 1-C CHS Ltd", "4.8"),
    ("INFINITY UNISEX SALON", "Salon", "Malad West", "+91 90499 06565", "", "", "BJ Patel Rd, opp. Raj Computer", "4.7"),
    ("THRIVE UNISEX SALOON", "Salon", "Malad West", "", "", "", "Shop no.6, Shah Arcade-2, Rani Sati Rd", "4.8"),
    ("Just Trim Salon", "Salon", "Malad West", "", "", "", "Shop no.5, Dheeraj Platinum, Chincholi Bunder Rd", "4.8"),

    # SALONS — BORIVALI
    ("Urbane Salon", "Salon", "Borivali West", "+91 91522 01957", "", "", "Shop no 2, Ekta Elegance, Link Road", "4.9"),
    ("Lookout Salon", "Salon", "Borivali West", "+91 98194 02381", "", "lookoutsalon.co.in", "Metro Pillar 247, Sun Sumit Enclave, New Link Rd", "4.8"),
    ("Ecoluxe Salon", "Salon", "Borivali West", "+91 86579 88918", "", "Instagram only", "Shop No 16, Satra Park, Shimpoli Rd", "4.8"),
    ("KS Beauty Lounge", "Salon", "Borivali West", "+91 91525 19145", "", "", "Shop No. 1-2, Kapoor Apartment, Chandavarkar Rd", "4.8"),
    ("Posh Salon & Spa", "Salon", "Borivali West", "+91 74003 74644", "", "", "Vastu Pinnacle, Eksar Rd", "4.8"),
    ("Lakme Salon SV Road", "Salon", "Borivali", "+91 93722 46874", "", "lakmesalon.in", "Shop No 5, 6/7, Swami Vivekanand Rd", "4.7"),
    ("Kapil's Salon", "Salon", "Borivali", "+91 86557 75644", "", "kapilssalon.com", "Shop No 42, Satra Park, Shimpoli Rd", "4.7"),
    ("Enrich Salon Satra Park", "Salon", "Borivali West", "+91 75065 75768", "", "enrichbeauty.com", "Shop No. 33 & 34, Satra Park, Shimpoli Rd", "4.7"),
    ("Lemon Salons Borivali", "Salon", "Borivali West", "+91 73044 53967", "", "Dingg booking", "Shop no 39, Satra Park, Shimpoli Rd", "4.5"),
    ("Enrich Salon Factory Lane", "Salon", "Borivali West", "+91 96198 52610", "", "enrichbeauty.com", "Shop No 12, Kent Garden Apartment, Factory Ln", "4.7"),
    ("Elinor Salon", "Salon", "Borivali West", "+91 93216 50585", "", "poonamgohel.in", "Shop no 001, Gokul Gagan Apartment, Factory Ln", "4.8"),
    ("Rac's Salon Borivali", "Salon", "Borivali West", "", "", "", "Shop No. 13, Kamla Rajesh Tower, Chandavarkar Rd", "4.9"),
    ("LYNN SALON", "Salon", "Borivali West", "", "", "", "Acropolis 1, Laxman Mhatre Rd", "4.8"),

    # GYMS — ANDHERI
    ("Waves Gym", "Gym", "Andheri West", "+91 22 6678 7970", "", "wavesgym.com", "5th Floor, Morya Estate, New Link Rd", "4.9"),
    ("IH Remind Fitness", "Gym", "Andheri East", "+91 79898 73329", "", "ihremindfitness.com", "Shop 101, Empress Business Bay, MIDC Central Rd", "4.7"),
    ("Plus Fitness 24/7", "Gym", "Andheri West", "+91 90820 92919", "", "plusfitness.co.in", "Salve Marg, 12A, Chandak Unicorn", "4.8"),
    ("Powerr Vibes Gym", "Gym", "Andheri", "+91 81040 04857", "", "Cult.fit listing", "Basement, Navkar Complex, opp. Andheri Court", "4.8"),
    ("GOODLIFE FITNESS", "Gym", "Andheri", "+91 72086 86233", "", "Instagram only", "Shop no 7, C Wing, Atul, Military Rd", "4.7"),
    ("i5 Fitness", "Gym", "Andheri East", "+91 97691 55315", "", "", "3rd Floor, T5, Pinnacle Business Park, Mahakali Caves Rd", "4.7"),
    ("Zion Fitness", "Gym", "Andheri East", "+91 77382 58050", "", "zionfitness.in", "H.E.K Compound, Mahakali Caves Rd", "4.6"),
    ("FORTUNE FITNESS", "Gym", "Andheri East", "+91 86929 60009", "", "fortunefitnessgym.com", "Nirvan Corporate, Pumphouse, opp. Aghadi Nagar", "4.4"),
    ("Fitness Factory", "Gym", "Andheri", "+91 96196 15000", "", "", "4th Floor, Idea Square, opp. Citi Mall", "4.1"),
    ("Stay Fit Andheri", "Gym", "Andheri West", "+91 77978 89788", "info@ironhouse.co.in", "ironhouse.co.in", "Kashibai Mahadev Marg, near The Club", "4.7"),
    ("FITNESS ART", "Gym", "Andheri", "+91 84229 96655", "", "", "1st floor, Deep Society, JP Rd, above DMart", "4.7"),
    ("Cloud 9 Fitness Club", "Gym", "Andheri West", "", "", "", "Paramount Bldg, New Link Rd", "4.3"),
    ("Planet Muscle Fitness", "Gym", "Andheri West", "", "", "", "Office 103, Ameya House, near Metro Azad Nagar", "4.8"),
    ("Fitness Hub", "Gym", "Andheri", "", "", "", "Shop 1/2, Teli Gali, opp. Criticare Hospital", "4.6"),

    # SPAS — LOWER PAREL
    ("Medis Luxury Spa", "Spa", "Lower Parel", "+91 83692 38697", "", "medisluxuryspaparel.in", "Shop 101, Raghuvanshi Mills, Senapati Bapat Marg", "5.0"),
    ("Nature Wellness Spa", "Spa", "Lower Parel", "+91 85917 62214", "", "naturespalowerparel.in", "3rd Floor, Tree Building, Raghuvanshi Mill Compound", "4.9"),
    ("Kyra Luxury Spa", "Spa", "Lower Parel", "+91 93728 31214", "", "kyraspalowerparel.com", "Shop 101, Times Tower, Kamla Mill Compound", "4.8"),
    ("Oceanic Spa", "Spa", "Lower Parel", "+91 81089 52829", "", "oceanicspalowerparel.in", "1st floor, Todi Mill Compound, Senapati Bapat Marg", "4.8"),
    ("Madisyn Spa", "Spa", "Lower Parel", "+91 90822 12203", "", "madisynspalowerparel.in", "1st Floor, Sumandar House, Raghuvanshi Mills", "4.8"),
    ("Mudraa Spa", "Spa", "Lower Parel", "+91 96196 08043", "", "mudraaspa.in", "Shop 102, Raghuvanshi Mills, Senapati Bapat Marg", "4.5"),
    ("Yuan Thai Spa", "Spa", "Lower Parel", "+91 98672 00557", "", "yuanthaispa.com", "1st Floor, Raghuvanshi Mills, Senapati Bapat Marg", "4.1"),
    ("R Spa", "Spa", "Lower Parel", "+91 89765 79019", "", "", "Ground Floor, Times Tower, Kamala Mills Compound", "4.8"),
    ("Spring Spa Worli", "Spa", "Worli", "+91 91371 06224", "", "", "Shop 209, Kachwala Building, Dr Annie Besant Rd", "4.8"),
    ("Richy Spa", "Spa", "Byculla", "+91 91676 14119", "", "richyspachinchpokli.com", "Office No. 6A-6C, Anand Estate, Sane Guruji Marg", "4.9"),
    ("Bodhi Luxury Spa", "Spa", "Lower Parel", "+91 96199 20630", "", "bodhigroup.in", "Store LG22, Phoenix Palladium Mall", "4.1"),
    ("St. Regis Spa", "Spa", "Lower Parel", "+91 88799 78888", "", "marriott.com", "Wellness Floor, Level 10, 462, Senapati Bapat Marg", "4.4"),
    ("Maitri Thai Spa", "Spa", "Lower Parel", "", "", "maitrispaparel.in", "Office No 101, Times Tower, Kamala Mills", "4.9"),
    ("Mudram Wellness Spa", "Spa", "Lower Parel", "", "", "", "1st Floor, Sumandar House, Raghuvanshi Mill Compound", "4.8"),

    # YOGA — BORIVALI
    ("Yog Shala Tat-Tvam-Asi", "Yoga", "Borivali", "+91 89762 15478", "", "", "Shop No. 2&3, Rajeshwari Building", "5.0"),
    ("Ashtanga Yog Shala", "Yoga", "Borivali", "+91 98923 59792", "", "", "1st floor, The Zone Mall, Chandavarkar Rd", "4.9"),
    ("Yoga365", "Yoga", "Borivali", "+91 98215 23064", "", "yoga365.in", "1st Floor, Plot no. 368, Besides Muthoot Finance", "4.9"),
    ("SoulYog Studio", "Yoga", "Borivali", "+91 83690 72930", "", "WhatsApp only", "Play and Learn Building, near Kanakia Aroha", "5.0"),
    ("Yog Kutir", "Yoga", "Borivali", "+91 92213 13717", "", "yogkutir.in", "1, Adinath Towers, Kanti Park Rd, Shimpoli", "4.8"),
    ("Rhythm & Flow Fitness", "Yoga", "Borivali", "+91 93269 47889", "", "rhythmandflow.in", "Abhiram CHSL, Amarkant Jha Marg", "4.9"),
    ("Iyengar Yoga Pratishthan", "Yoga", "Borivali", "+91 98733 94133", "", "", "64, D1, Dwakesh Park, above HDFC Bank", "4.9"),
    ("Flow & Sway Fitness", "Yoga", "Borivali", "+91 87793 79503", "", "", "Clover Grove, Shop 19 & 20, Chikoowadi", "4.9"),
    ("SHAPE IN YOGA", "Yoga", "Borivali", "+91 77383 00448", "", "", "27, RSC Rd Number 26, near St. Rock's School", "5.0"),
    ("OM YOGA STUDIO", "Yoga", "Borivali", "+91 88796 04929", "", "", "Tata Power House, RSM Gym, Phulpakharu Garden", "4.9"),
    ("Yoga with Fennil", "Yoga", "Borivali", "", "", "", "C-1, Hariom Apartment, Swami Vivekanand Rd", "4.9"),
    ("Shushant Dhuri Yoga", "Yoga", "Vile Parle East", "080 6987 4345", "", "", "Vile Parle East, 400057", "5.6"),
    ("K Fitness & Yoga", "Yoga", "Ghatkopar West", "080 3542 8557", "", "", "Ghatkopar West, 400084", "4.9"),
    ("Nivedita Dhage Yoga", "Yoga", "Dadar West", "080 6983 1607", "", "", "Dadar West, 400028", "6.7"),

    # PET GROOMING
    ("Paw Lovey Spa", "Pet Grooming", "Andheri", "+91 75068 80515", "", "", "Shop no.11, Takshila 29B, Mahakali Caves Rd", "5.0"),
    ("Cuddle Grooms", "Pet Grooming", "Mumbai", "+91 85911 89518", "", "", "Shop no.22, Pink Corner", "5.0"),
    ("Pampurr Pets Grooming", "Pet Grooming", "Andheri", "+91 98193 99281", "", "", "Shop No A-2, Crystal Colony", "4.9"),
    ("Furry Foots Pet Grooming", "Pet Grooming", "Magathane", "+91 91369 30520", "", "", "Vasant Marvel Complex, near Magathane Metro", "4.9"),
    ("We Love Pets Grooming", "Pet Grooming", "Mumbai", "+91 83695 99051", "", "", "Shop No 9, Nilkhanth Soc.", "4.8"),
    ("Zane's Pet Spa", "Pet Grooming", "Lower Parel", "+91 91524 92637", "", "", "1st Floor, Todi Mill Compound", "4.8"),
    ("FUR N TAILS", "Pet Grooming", "Mumbai", "+91 70211 56081", "", "", "Shop No 2, Vasuki Building", "4.7"),
    ("The Happy Tails", "Pet Grooming", "Mumbai", "+91 77158 88881", "", "", "Prabha Niketan, Plot No. 338", "4.6"),

    # PHYSIOTHERAPY
    ("Movmed Physiotherapy", "Physio", "Andheri West", "080 4810 4585", "", "", "Andheri West, Mumbai 400053", "7.3"),
    ("Dr Sahil's Physiotherapy", "Physio", "Thane West", "+91 90297 31346", "", "advphysiotherapyclinic.com", "Shop 13, Siddhachal Phase 8, Pokharan Rd", "6.7"),
    ("Inest Family Clinic", "Physio", "Panvel", "+91 79779 08975", "", "vedantaaclinic.com", "Shop no 2, Govind Nilayam, opp. Status hotel", "7.1"),
    ("EVOLUTION OF MOVEMENT", "Physio", "Colaba", "080 4812 1136", "", "", "Colaba, 400005", "3.9"),

    # DRIVING SCHOOLS
    ("Superdrive Zone", "Driving School", "Kandivali West", "080 4810 2575", "", "", "Kandivali West, 400067", "5.0"),
    ("Vaidya Motor Training", "Driving School", "Girgaon", "+91 98216 27774", "", "", "Khotachiwadi Rd, Charni Road East", "5.0"),
    ("Toyota Driving School", "Driving School", "Malad West", "+91 77100 33877", "", "toyotabharat.com", "504, New Link Rd, opp. Goregaon Sports Club", "5.0"),
]

# ── HEADERS ──
HEADERS = ["Sr No", "Business Name", "Category", "Area", "Phone Number", "Email", "Website", "Address", "Rating"]
COL_WIDTHS = [6, 32, 16, 18, 22, 26, 28, 48, 8]

for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# ── DATA ROWS ──
for i, lead in enumerate(leads):
    row = i + 2
    name, cat, area, phone, email, website, addr, rating = lead
    vals = [i + 1, name, cat, area, phone, email, website, addr, rating]
    fill = alt_fill if i % 2 == 0 else white_fill

    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = phone_font if col == 5 else data_font
        c.fill = fill
        c.border = thin
        c.alignment = Alignment(vertical='center', horizontal='center' if col in (1, 9) else 'left')
    ws.row_dimensions[row].height = 20

# ── COLUMN WIDTHS ──
for idx, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(idx)].width = w

# ── AUTO FILTER ──
ws.auto_filter.ref = f"A1:I{1 + len(leads)}"

# ── SAVE TO DESKTOP ──
path = "/Users/bhaveshashokwaghmare/Desktop/Mumbai_Leads.xlsx"
wb.save(path)
print(f"✅ Saved: {path}")
print(f"   Total leads: {len(leads)}")
print(f"   With phone: {len([l for l in leads if l[3]])}")
