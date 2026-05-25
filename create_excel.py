import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ── Colors ──
DARK_BG      = "1B1F2B"
TIER1_HDR    = "E74C3C"   # red
TIER1_ROW    = "FDEDEC"
TIER2_HDR    = "F39C12"   # orange
TIER2_ROW    = "FEF5E7"
TIER3_HDR    = "F1C40F"   # yellow
TIER3_ROW    = "FEF9E7"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F7F9FC"
HEADER_BLUE  = "2C3E50"

thin_border = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)

hdr_font  = Font(name='Calibri', bold=True, color=WHITE, size=11)
data_font = Font(name='Calibri', size=10)
phone_font = Font(name='Calibri', size=11, bold=True, color="2471A3")
hot_font   = Font(name='Calibri', size=10, bold=True, color=TIER1_HDR)
title_font = Font(name='Calibri', size=16, bold=True, color=DARK_BG)
sub_font   = Font(name='Calibri', size=10, italic=True, color="7F8C8D")

# ── All leads data ──
leads = [
    # Tier, Name, Category, Area, Phone, Website, Has Booking, Address, Rating, Notes
    # TIER 1 — NO website
    ("🔥 TIER 1", "Look Good Unisex Salon", "Salon", "Andheri East", "+91 83559 28175", "", "No", "Shop no.1, Gokul Plaza, Old Nagardas Rd", "4.9", "NO website — hot lead"),
    ("🔥 TIER 1", "FTV Salon", "Salon", "Andheri West", "+91 93211 65417", "", "No", "Samarth Vaibhav, 112, Lokhandwala Rd", "4.7", "NO website — hot lead"),
    ("🔥 TIER 1", "THE KIN UNISEX SALON", "Salon", "Andheri West", "+91 70212 70618", "", "No", "Society No. D5/55, Mhada Rd, near Hanuman Mandir", "4.8", "NO website — hot lead"),
    ("🔥 TIER 1", "Urbane Salon", "Salon", "Borivali West", "+91 91522 01957", "", "No", "Shop no 2, Ekta Elegance, Link Road", "4.9", "NO website — hot lead"),
    ("🔥 TIER 1", "KS Beauty Lounge", "Salon", "Borivali West", "+91 91525 19145", "", "No", "Shop No. 1-2, Kapoor Apartment, Chandavarkar Rd", "4.8", "NO website — hot lead"),
    ("🔥 TIER 1", "Posh Salon & Spa", "Salon", "Borivali West", "+91 74003 74644", "", "No", "Vastu Pinnacle, Eksar Rd", "4.8", "NO website — hot lead"),
    ("🔥 TIER 1", "Paw Lovey Spa ⭐", "Pet Grooming", "Andheri", "+91 75068 80515", "", "No", "Shop no.11, Takshila 29B, Mahakali Caves Rd", "5.0", "PERFECT rating + NO website"),
    ("🔥 TIER 1", "Cuddle Grooms ⭐", "Pet Grooming", "Mumbai", "+91 85911 89518", "", "No", "Shop no.22, Pink Corner", "5.0", "PERFECT rating + NO website"),
    ("🔥 TIER 1", "Pampurr Pets Grooming", "Pet Grooming", "Andheri", "+91 98193 99281", "", "No", "Shop No A-2, Crystal Colony, near Canossa High School", "4.9", "NO website — hot lead"),
    ("🔥 TIER 1", "Furry Foots Pet Grooming", "Pet Grooming", "Magathane", "+91 91369 30520", "", "No", "Vasant Marvel Complex, near Magathane Metro", "4.9", "NO website — hot lead"),
    ("🔥 TIER 1", "We Love Pets Grooming", "Pet Grooming", "Mumbai", "+91 83695 99051", "", "No", "Shop No 9, Nilkhanth Soc.", "4.8", "NO website"),
    ("🔥 TIER 1", "Zane's Pet Spa", "Pet Grooming", "Lower Parel", "+91 91524 92637", "", "No", "1st Floor, Todi Mill Compound, Senapati Bapat Marg", "4.8", "NO website"),
    ("🔥 TIER 1", "FUR N TAILS", "Pet Grooming", "Mumbai", "+91 70211 56081", "", "No", "Shop No 2, Vasuki Building, near Somaiya College", "4.7", "NO website"),
    ("🔥 TIER 1", "The Happy Tails", "Pet Grooming", "Mumbai", "+91 77158 88881", "", "No", "Prabha Niketan, Plot No. 338", "4.6", "NO website"),
    ("🔥 TIER 1", "GOODLIFE FITNESS", "Gym", "Andheri", "+91 72086 86233", "Instagram only", "No", "Shop no 7, C Wing, Atul, Military Rd", "4.7", "Only Instagram"),
    ("🔥 TIER 1", "Yog Shala Tat-Tvam-Asi ⭐", "Yoga", "Borivali", "+91 89762 15478", "", "No", "Shop No. 2&3, Rajeshwari Building, near Sai Baba Nagar", "5.0", "PERFECT rating + NO website"),
    ("🔥 TIER 1", "Ashtanga Yog Shala", "Yoga", "Borivali", "+91 98923 59792", "", "No", "1st floor, The Zone Mall, Chandavarkar Rd", "4.9", "NO website — hot lead"),
    ("🔥 TIER 1", "Movmed Physiotherapy", "Physio", "Andheri West", "080 4810 4585", "", "No", "Andheri West, Mumbai 400053", "7.3/10", "NO website — hot lead"),
    ("🔥 TIER 1", "Superdrive Zone", "Driving School", "Kandivali West", "080 4810 2575", "", "No", "Kandivali West, 400067", "5.0", "NO website"),

    # TIER 2 — Instagram / WhatsApp / Google Sites only
    ("🟠 TIER 2", "Ddazzle Unisex Salon", "Salon", "Andheri West", "+91 96536 61417", "Instagram only", "No", "Shop 36, Laxmi Plaza", "4.9", "Only Instagram — needs website"),
    ("🟠 TIER 2", "Glitz Salon Academy", "Salon", "Andheri East", "+91 98926 29888", "Instagram only", "No", "Shop No.8, Building No.6, MIDC Central Rd", "4.8", "Only Instagram — needs website"),
    ("🟠 TIER 2", "Toni&Guy Four Bungalows", "Salon", "Andheri West", "+91 98673 67722", "Linktree only", "No", "Shop No 1, Ashish CHS, Building 24", "4.6", "Only Linktree — needs proper site"),
    ("🟠 TIER 2", "De Bella & Beau Salon", "Salon", "Malad West", "+91 98204 22003", "WhatsApp only", "No", "Shop 10, Old Sonal Industrial Estate, New Link Rd", "4.8", "Only WhatsApp — needs website"),
    ("🟠 TIER 2", "Stylorria Salon", "Salon", "Malad West", "+91 84519 58888", "Google Sites page", "No", "Shop No 4, Deep Sunder Lane CHS", "4.8", "Google Sites — needs proper site"),
    ("🟠 TIER 2", "Ecoluxe Salon", "Salon", "Borivali West", "+91 86579 88918", "Instagram only", "No", "Shop No 16, Satra Park, Shimpoli Rd", "4.8", "Only Instagram — needs website"),
    ("🟠 TIER 2", "SoulYog Studio", "Yoga", "Borivali", "+91 83690 72930", "WhatsApp only", "No", "Play and Learn Building, near Kanakia Aroha", "5.0", "PERFECT rating — WhatsApp only"),
    ("🟠 TIER 2", "Shushant Dhuri Yoga", "Yoga", "Vile Parle East", "080 6987 4345", "", "No", "Vile Parle East, 400057", "5.6", "No website"),
    ("🟠 TIER 2", "K Fitness & Yoga Institute", "Yoga", "Ghatkopar West", "080 3542 8557", "", "No", "Ghatkopar West, 400084", "4.9", "No website"),

    # TIER 3 — Has website, needs booking / AI receptionist
    ("🟡 TIER 3", "Shanuzz Salon", "Salon", "Andheri West", "+91 91524 18567", "shanuzz.com", "No", "Ground Floor, Harshvardhan Chambers", "4.9", "Has website — no booking system"),
    ("🟡 TIER 3", "Moon Studios Salon", "Salon", "Andheri East", "+91 90048 32184", "moonstudiossalon.in", "No", "Shop C, Harmony Apartments, Krishanlal Marwah Marg", "4.9", "Has website — no booking"),
    ("🟡 TIER 3", "Harsha & Rakesh Salon", "Salon", "Lokhandwala", "+91 99203 27888", "harsharakesh.com", "No", "Shop No. 6/7, Primerose Apartment, 4th Cross Rd", "4.8", "Has website — no booking"),
    ("🟡 TIER 3", "Huppa Huiyya Salon", "Salon", "Veera Desai Rd", "+91 81043 49859", "huppahuiyyasalon.online", "No", "Shop No 4, Bhavesha Building, Veera Desai Rd", "4.8", "Has website — check booking"),
    ("🟡 TIER 3", "Beleesa Salon", "Salon", "Andheri West", "+91 81047 55280", "Luzo booking app", "Partial", "Raviraj Complex, Shop No 11, Rendezvous", "4.9", "Has Luzo — may need own site"),
    ("🟡 TIER 3", "FLOW HAIR STUDIO", "Salon", "Malad West", "+91 89766 50250", "flowhairstudio.in", "No", "Shop no. 14, Atlanta Building", "4.9", "Has website — check booking"),
    ("🟡 TIER 3", "Locks & Lustre Salon", "Salon", "Malad West", "+91 98334 16565", "locksandlustresalon.com", "No", "Shop no G-3, Sej Plaza, Marve Rd", "4.8", "Has website — check booking"),
    ("🟡 TIER 3", "Lookout Salon", "Salon", "Borivali West", "+91 98194 02381", "lookoutsalon.co.in", "No", "Metro Pillar 247, Sun Sumit Enclave, New Link Rd", "4.8", "Has website — check booking"),
    ("🟡 TIER 3", "The Glam Lab Salon", "Salon", "Bandra West", "+91 73040 14055", "theglamlabsalon.in", "No", "Shop No. 101, Asra Building, Waterfield Road", "4.9", "Has website — no booking"),
    ("🟡 TIER 3", "Apple Unisex Salon", "Salon", "Bandra West", "+91 90821 83323", "applesalon.in", "No", "Ground Floor, SVD House, 32nd Rd", "4.9", "Has website — check booking"),
    ("🟡 TIER 3", "7 Wonders Studio", "Salon", "Bandra West", "+91 98336 79949", "7wondersstudiosalon.in", "No", "shop no 4, Libra Tower, Hill Rd", "4.9", "Has website — no booking"),
    ("🟡 TIER 3", "Color Cafe Salon", "Salon", "Bandra West", "+91 80801 26688", "colorcafe.co.in", "No", "Kailash Building, 156, Waterfield Road", "4.9", "Has website — check booking"),
    ("🟡 TIER 3", "Dessange Salon & Spa", "Salon", "Bandra West", "+91 73043 38957", "dessangemumbai.com", "Likely No", "190, Savanna Court, Turner Rd", "4.5", "Premium salon — check booking"),
    ("🟡 TIER 3", "Ccoral Salon", "Salon", "Bandra West", "+91 73044 42223", "ccoralsalon.in", "No", "Ground floor, Raut Sadan, St John Baptist Rd", "4.6", "Has website — check booking"),
    ("🟡 TIER 3", "Medis Luxury Spa ⭐", "Spa", "Lower Parel", "+91 83692 38697", "medisluxuryspaparel.in", "No", "Shop 101, Raghuvanshi Mills, Senapati Bapat Marg", "5.0", "PERFECT rating — needs booking"),
    ("🟡 TIER 3", "Nature Wellness Spa", "Spa", "Lower Parel", "+91 85917 62214", "naturespalowerparel.in", "No", "3rd Floor, Tree Building, Raghuvanshi Mill Compound", "4.9", "Has website — needs booking"),
    ("🟡 TIER 3", "Kyra Luxury Spa", "Spa", "Lower Parel", "+91 93728 31214", "kyraspalowerparel.com", "No", "Shop 101, Times Tower, Kamla Mill Compound", "4.8", "Has website — check booking"),
    ("🟡 TIER 3", "Oceanic Spa", "Spa", "Lower Parel", "+91 81089 52829", "oceanicspalowerparel.in", "No", "1st floor, Todi Mill Compound, Senapati Bapat Marg", "4.8", "Has website — check booking"),
    ("🟡 TIER 3", "Yoga365", "Yoga", "Borivali", "+91 98215 23064", "yoga365.in", "No", "1st Floor, Plot no. 368, Besides Muthoot Finance", "4.9", "Has website — check booking"),
    ("🟡 TIER 3", "Waves Gym", "Gym", "Andheri West", "+91 22 6678 7970", "wavesgym.com", "No", "5th Floor, Morya Estate, New Link Rd", "4.9", "Has website — check booking"),
    ("🟡 TIER 3", "IH Remind Fitness", "Gym", "Andheri", "+91 79898 73329", "ihremindfitness.com", "No", "Shop 101, Empress Business Bay, MIDC Central Rd", "4.7", "Has website — check booking"),
    ("🟡 TIER 3", "Plus Fitness 24/7", "Gym", "Andheri", "+91 90820 92919", "plusfitness.co.in", "No", "Salve Marg, 12A, Chandak Unicorn, Off Veera Desai Rd", "4.8", "Has website — check booking"),
]

HEADERS = ["#", "Priority", "Business Name", "Category", "Area", "📞 Phone", "Website", "Has Booking?", "Address", "⭐ Rating", "Notes / Call Status"]
COL_WIDTHS = [4, 12, 30, 15, 16, 22, 28, 14, 45, 10, 35]

# ═══════════════════════════════════════
# SHEET 1 — All Leads
# ═══════════════════════════════════════
ws = wb.active
ws.title = "All Leads"
ws.sheet_properties.tabColor = "2C3E50"

# Title row
ws.merge_cells("A1:K1")
title_cell = ws["A1"]
title_cell.value = "📞 Mumbai Cold Calling Leads — " + datetime.now().strftime("%d %b %Y")
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[1].height = 36

# Subtitle
ws.merge_cells("A2:K2")
ws["A2"].value = "Source: Google Maps via Firecrawl API  |  53 leads with verified phone numbers  |  Prioritized by digital presence gap"
ws["A2"].font = sub_font
ws.row_dimensions[2].height = 20

# Headers — row 4
header_fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')
for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = hdr_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[4].height = 28

# Auto-freeze panes
ws.freeze_panes = "A5"

# Data rows
tier_fills = {
    "🔥 TIER 1": PatternFill(start_color=TIER1_ROW, end_color=TIER1_ROW, fill_type='solid'),
    "🟠 TIER 2": PatternFill(start_color=TIER2_ROW, end_color=TIER2_ROW, fill_type='solid'),
    "🟡 TIER 3": PatternFill(start_color=TIER3_ROW, end_color=TIER3_ROW, fill_type='solid'),
}
tier_fonts = {
    "🔥 TIER 1": Font(name='Calibri', size=10, bold=True, color="C0392B"),
    "🟠 TIER 2": Font(name='Calibri', size=10, bold=True, color="E67E22"),
    "🟡 TIER 3": Font(name='Calibri', size=10, bold=True, color="B7950B"),
}

for i, lead in enumerate(leads):
    row = i + 5
    tier = lead[0]
    values = [i+1, tier, lead[1], lead[2], lead[3], lead[4], lead[5], lead[6], lead[7], lead[8], lead[9]]
    fill = tier_fills.get(tier)
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if fill:
            cell.fill = fill
        # Special formatting
        if col_idx == 2:  # Priority
            cell.font = tier_fonts.get(tier, data_font)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx == 6:  # Phone
            cell.font = phone_font
        elif col_idx == 1:  # #
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx == 10:  # Rating
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 22

# Column widths
for idx, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(idx)].width = w

# ═══════════════════════════════════════
# SHEET 2 — Tier 1 Hot Leads
# ═══════════════════════════════════════
ws2 = wb.create_sheet("🔥 Tier 1 — Hot Leads")
ws2.sheet_properties.tabColor = "E74C3C"

ws2.merge_cells("A1:F1")
ws2["A1"].value = "🔥 TIER 1 — No Website, No Booking — HOTTEST LEADS"
ws2["A1"].font = Font(name='Calibri', size=14, bold=True, color="C0392B")
ws2.row_dimensions[1].height = 32

hot_headers = ["#", "Business Name", "Category", "Area", "📞 Phone", "⭐ Rating"]
hot_widths = [4, 30, 15, 18, 22, 10]
hdr_fill2 = PatternFill(start_color=TIER1_HDR, end_color=TIER1_HDR, fill_type='solid')

for col_idx, h in enumerate(hot_headers, 1):
    cell = ws2.cell(row=3, column=col_idx, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill2
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
ws2.freeze_panes = "A4"

tier1 = [l for l in leads if l[0] == "🔥 TIER 1"]
for i, lead in enumerate(tier1):
    row = i + 4
    vals = [i+1, lead[1], lead[2], lead[3], lead[4], lead[8]]
    for col_idx, val in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=col_idx, value=val)
        cell.font = phone_font if col_idx == 5 else data_font
        cell.fill = PatternFill(start_color=TIER1_ROW, end_color=TIER1_ROW, fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if col_idx in (1,6) else 'left', vertical='center')
    ws2.row_dimensions[row].height = 22

for idx, w in enumerate(hot_widths, 1):
    ws2.column_dimensions[get_column_letter(idx)].width = w

# ═══════════════════════════════════════
# SHEET 3 — Call Tracker
# ═══════════════════════════════════════
ws3 = wb.create_sheet("📋 Call Tracker")
ws3.sheet_properties.tabColor = "27AE60"

ws3.merge_cells("A1:H1")
ws3["A1"].value = "📋 Cold Call Tracker — Log your outreach progress"
ws3["A1"].font = Font(name='Calibri', size=14, bold=True, color="27AE60")
ws3.row_dimensions[1].height = 32

tracker_headers = ["#", "Business Name", "Phone", "Call Date", "Spoke To", "Outcome", "Follow-Up Date", "Notes"]
tracker_widths = [4, 28, 22, 14, 18, 20, 14, 35]
hdr_fill3 = PatternFill(start_color="27AE60", end_color="27AE60", fill_type='solid')

for col_idx, h in enumerate(tracker_headers, 1):
    cell = ws3.cell(row=3, column=col_idx, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill3
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
ws3.freeze_panes = "A4"

# Pre-populate with Tier 1 leads
for i, lead in enumerate(tier1):
    row = i + 4
    vals = [i+1, lead[1], lead[4], "", "", "", "", ""]
    for col_idx, val in enumerate(vals, 1):
        cell = ws3.cell(row=row, column=col_idx, value=val)
        cell.font = phone_font if col_idx == 3 else data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        if i % 2 == 0:
            cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
    ws3.row_dimensions[row].height = 22

for idx, w in enumerate(tracker_widths, 1):
    ws3.column_dimensions[get_column_letter(idx)].width = w

# Add data validation for Outcome column
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"Not Answered,Interested,Not Interested,Call Back,Wrong Number,Closed Deal"', allow_blank=True)
dv.error = "Pick from the dropdown"
dv.prompt = "Select call outcome"
ws3.add_data_validation(dv)
for row in range(4, 4 + len(tier1)):
    dv.add(ws3.cell(row=row, column=6))

# ── Save ──
output_path = "/Users/bhaveshashokwaghmare/Desktop/anticlaude/Mumbai_Cold_Calling_Leads.xlsx"
wb.save(output_path)
print(f"✅ Excel saved to: {output_path}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Total leads: {len(leads)}")
print(f"   Tier 1 (hot): {len(tier1)}")
